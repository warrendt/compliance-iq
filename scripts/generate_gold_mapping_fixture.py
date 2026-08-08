#!/usr/bin/env python3
"""Derive the NCSP v2.0 gold-mapping regression fixture from the expert workbook.

The workbook (``Mapping_Example_Perfect.xlsx``) is a hand-built, expert-reviewed
mapping of the UAE NCSP v2.0 control set onto Azure Policy. It is the quality bar
the mapping engine is measured against, so sheet ``NCSP v2.0 Controls`` is
extracted here into a plain JSON fixture that tests can consume without needing
the binary workbook (or ``openpyxl``) in the test path.

Two normalisations are applied on the way out:

* The customer's trading name in the ``Responsibility`` column is replaced with
  the neutral token ``Customer`` so no customer identity is committed to the repo.
* Policy definition GUIDs are parsed out of the free-text ``Policy Definition ID``
  column into a list, lower-cased, so recall can be measured set-wise.

Usage::

    python scripts/generate_gold_mapping_fixture.py <workbook.xlsx> [-o <out.json>]

The fixture is committed, so this only needs re-running when the workbook changes.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List

# Sheet layout. The header sits on row 5 and data runs to row 142 (137 controls).
SHEET_NAME = "NCSP v2.0 Controls"
HEADER_ROW = 5
FIRST_DATA_ROW = 6
LAST_DATA_ROW = 142
EXPECTED_CONTROL_COUNT = 137

LEGEND_SHEET = "Legend"

_GUID_RE = re.compile(
    r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}"
)

# The workbook names the customer directly (in the Responsibility column and
# throughout the free-text columns). Every occurrence is replaced with a neutral
# token so no customer identity is committed to the repo. "Microsoft" is left
# as-is since it names the CSP, not a customer.
_NEUTRAL_RESPONSIBILITY = "Customer"
_KNOWN_CSP_RESPONSIBILITY = "microsoft"

# Customer trading name as it appears in the workbook, with the possessive and
# sentence-initial capitalised forms. Ordered longest-first so the possessive is
# consumed before the bare form. Applied to every extracted string.
_CUSTOMER_NAME_SUBSTITUTIONS = (
    ("e&'s", "the Customer's"),
    ("E&'s", "the Customer's"),
    ("e&", "the Customer"),
    ("E&", "the Customer"),
)


def _scrub_customer_name(text: str) -> str:
    """Replace the customer's trading name with a neutral token.

    Applied to every extracted cell. ``e&`` is not a word-boundary-friendly
    token (``&`` is non-word), so plain ordered string replacement is used
    rather than a regex with ``\\b``.
    """
    for name, replacement in _CUSTOMER_NAME_SUBSTITUTIONS:
        text = text.replace(name, replacement)
    return text

# Coverage categories as spelled in the workbook, mapped to the stable internal
# constants in ``app.services.coverage``. Compared case-folded and with the
# workbook's non-ASCII dashes normalised.
COVERAGE_CATEGORY_MAP = {
    "azure policy enforced": "A_AzurePolicy",
    "azure/entra config - partial": "B_AzureConfig",
    "microsoft attested": "D_MicrosoftAttestation",
    "process / organisational": "C_Process",
}

# Enforcement planes as spelled in the workbook. Retained as documentation of the
# permitted vocabulary the mapping engine must reproduce.
ENFORCEMENT_PLANES = (
    "SLZ (deploy-time)",
    "Defender (run-time)",
    "SLZ (deploy-time) + Defender (run-time)",
    "None (manual control)",
)


def _clean(value: Any) -> str:
    """Normalise a cell to a stripped string with ASCII-safe punctuation.

    The workbook was authored in Excel on Windows and carries cp1252 smart
    quotes, en/em dashes and non-breaking spaces. Those are folded to ASCII so
    the committed fixture is diff-friendly and stable across platforms.
    """
    if value is None:
        return ""
    text = str(value)
    for bad, good in (
        ("\u2018", "'"), ("\u2019", "'"),
        ("\u201c", '"'), ("\u201d", '"'),
        ("\u2013", "-"), ("\u2014", "-"),
        ("\u2026", "..."), ("\xa0", " "),
    ):
        text = text.replace(bad, good)
    return _scrub_customer_name(" ".join(text.split()))


def _normalise_responsibility(value: str) -> str:
    """Map the workbook's responsibility values onto Customer / Microsoft."""
    if value.strip().casefold() == _KNOWN_CSP_RESPONSIBILITY:
        return "Microsoft"
    return _NEUTRAL_RESPONSIBILITY if value.strip() else ""


def _normalise_coverage(value: str) -> str:
    """Map a workbook coverage category onto the internal constant.

    Unknown values raise: a silently mis-bucketed category would corrupt the
    classification harness, and the workbook only has four.
    """
    key = _clean(value).casefold()
    if key not in COVERAGE_CATEGORY_MAP:
        raise ValueError(f"Unrecognised coverage category: {value!r}")
    return COVERAGE_CATEGORY_MAP[key]


def _split_effects(value: str) -> List[str]:
    """Split the ``Effect`` cell into individual effects, dropping N/A."""
    text = _clean(value)
    if not text or text.casefold() == "n/a":
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def _split_list(value: str) -> List[str]:
    """Split a semicolon/comma separated cell, dropping N/A and None markers."""
    text = _clean(value)
    if not text or text.casefold() in {"n/a", "none", "-"}:
        return []
    parts = re.split(r"[;,]", text)
    return [p.strip() for p in parts if p.strip() and p.strip().casefold() != "none"]


def extract_controls(worksheet) -> List[Dict[str, Any]]:
    """Pull the 137 control rows out of the gold sheet.

    Pure with respect to the worksheet (read-only), so it is unit-testable
    against a synthetic sheet.
    """
    header = [_clean(cell.value) for cell in worksheet[HEADER_ROW]]

    def col(name: str) -> int:
        return header.index(name) + 1

    controls: List[Dict[str, Any]] = []
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        get = lambda name: _clean(worksheet.cell(row, col(name)).value)  # noqa: E731

        control_id = get("ID")
        if not control_id:
            continue

        raw_policy_ids = get("Policy Definition ID")
        controls.append(
            {
                "control_id": control_id,
                "section": get("Section"),
                "domain": get("Domain"),
                "control_name": get("Control Name"),
                "responsibility": _normalise_responsibility(get("Responsibility")),
                "description": get("Description"),
                "how_to_meet": get("How to meet the control objectives"),
                "reason": get("Reason"),
                "policy_definition_ids": [
                    g.lower() for g in _GUID_RE.findall(raw_policy_ids)
                ],
                "policy_type": get("Policy Type"),
                "defender_recommendations": _split_list(
                    get("Defender for Cloud Recommendation")
                ),
                "mcsb_control_ids": _split_list(get("MCSB Control ID")),
                "enforcement_plane": get("Enforcement Plane"),
                "effects": _split_effects(get("Effect")),
                "coverage_category": _normalise_coverage(get("Coverage Category")),
            }
        )
    return controls


def extract_legend(worksheet) -> Dict[str, str]:
    """Pull the coverage-category definitions out of the Legend sheet.

    These are the definitions the classification prompt is grounded in, so they
    travel with the fixture rather than being paraphrased in the prompt.
    """
    legend: Dict[str, str] = {}
    for row in worksheet.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        key, value = _clean(row[0]), _clean(row[1])
        if not key or not value:
            continue
        mapped = COVERAGE_CATEGORY_MAP.get(key.casefold())
        if mapped:
            legend[mapped] = value
    return legend


def build_fixture(workbook) -> Dict[str, Any]:
    """Assemble the full fixture payload from an open workbook."""
    controls = extract_controls(workbook[SHEET_NAME])
    if len(controls) != EXPECTED_CONTROL_COUNT:
        raise ValueError(
            f"Expected {EXPECTED_CONTROL_COUNT} controls, extracted {len(controls)}"
        )

    counts: Dict[str, int] = {}
    for control in controls:
        counts[control["coverage_category"]] = (
            counts.get(control["coverage_category"], 0) + 1
        )

    return {
        "framework": "UAE NCSP v2.0",
        "source": "expert-reviewed mapping workbook (Mapping_Example_Perfect.xlsx)",
        "control_count": len(controls),
        "coverage_counts": counts,
        "coverage_definitions": extract_legend(workbook[LEGEND_SHEET]),
        "controls": controls,
    }


def main(argv: List[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Path to the gold mapping .xlsx")
    parser.add_argument(
        "-o",
        "--output",
        default=str(
            Path(__file__).resolve().parent.parent
            / "app" / "tests" / "fixtures" / "ncsp_v2_gold_mapping.json"
        ),
        help="Destination JSON fixture path",
    )
    args = parser.parse_args(argv)

    import openpyxl

    workbook = openpyxl.load_workbook(args.workbook, data_only=True)
    fixture = build_fixture(workbook)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(fixture, indent=2, ensure_ascii=True) + "\n", encoding="utf-8"
    )

    print(f"Wrote {fixture['control_count']} controls to {out_path}")
    print(f"Coverage counts: {fixture['coverage_counts']}")
    guids = {g for c in fixture["controls"] for g in c["policy_definition_ids"]}
    print(
        f"Controls with policy GUIDs: "
        f"{sum(1 for c in fixture['controls'] if c['policy_definition_ids'])}, "
        f"distinct GUIDs: {len(guids)}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
